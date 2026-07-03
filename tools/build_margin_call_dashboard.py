#!/usr/bin/env python3
"""
Build a Securities-Backed Lending Margin Call dashboard (DBS credit methodology).

Methodology implemented (from DBS Credit Refresher training):
  CV per position           = Quantity x Price x LV%           (slide "Collateral Value")
  Concentration haircut     = 4 dimensions (Single Counter, Issuer, Country of Risk,
                              Asset Type). Conc Ratio = group CV / Total CV for
                              Concentration; Adj Factor = threshold / ratio when
                              breached (else 1); Conc Adj CV = CV x MIN(factors);
                              denominator = Marketable CV + MIN(Cash CV, cap% x
                              Marketable CV).                   (slide "Concentration Haircut")
  HCFX                      = 10% haircut on residual CV in currency buckets that is
                              allocated to exposure in a different currency; USD & HKD
                              form one bucket (no HCFX between them).
                                                                (slide "Currency Mismatch Hair-cut")
  Net Collateral Value (CV2)= Conc Adj CV - HCFX; margin call when
                              Exposure > CV2 x trigger.

Nothing is hardcoded in the calculation chain: every threshold, LV, FX rate,
bucket and facility lives on the Config sheet and the engine recalculates for any
portfolio pasted into the Positions sheet.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

OUT = "/home/user/Xx/Margin_Call_Dashboard.xlsx"

N_ROWS = 200            # position capacity
R0 = 5                  # first data row on Positions
R1 = R0 + N_ROWS - 1    # 204
TOT = R1 + 1            # 205 totals row

# ---------------------------------------------------------------- styles
NAVY   = "1F3864"; BLUE = "2E5AA8"; LGREY = "F2F2F2"; MGREY = "D9D9D9"
GOOD   = "C6EFCE"; GOODF = "006100"
BAD    = "FFC7CE"; BADF  = "9C0006"
WARN   = "FFEB9C"; WARNF = "9C6500"
INPUTC = "FFF2CC"   # user-input cells
CALCC  = "DDEBF7"   # key calculated cells

f_title  = Font(bold=True, size=14, color="FFFFFF")
f_sect   = Font(bold=True, size=11, color="FFFFFF")
f_hdr    = Font(bold=True, size=9, color="FFFFFF")
f_lbl    = Font(size=10)
f_lblb   = Font(bold=True, size=10)
f_small  = Font(size=8, italic=True, color="595959")
f_kpi    = Font(bold=True, size=11)

fill_title = PatternFill("solid", fgColor=NAVY)
fill_sect  = PatternFill("solid", fgColor=BLUE)
fill_hdr   = PatternFill("solid", fgColor=NAVY)
fill_in    = PatternFill("solid", fgColor=INPUTC)
fill_calc  = PatternFill("solid", fgColor=CALCC)
fill_grey  = PatternFill("solid", fgColor=LGREY)
fill_tot   = PatternFill("solid", fgColor=MGREY)

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '#,##0'
MONEY2 = '#,##0.00'
PCT = '0.0%'
PCT2 = '0.00%'
FACT = '0.000'

wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = openpyxl.Workbook()

def name(nm, ref):
    wb.defined_names.add(DefinedName(nm, attr_text=ref))

def hdr_row(ws, row, headers, start_col=1, fill=fill_hdr, font=f_hdr):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = font; c.fill = fill; c.alignment = center; c.border = border

def sect(ws, row, text, span):
    ws.cell(row=row, column=1, value=text).font = f_sect
    ws.cell(row=row, column=1).fill = fill_sect
    for col in range(2, span + 1):
        ws.cell(row=row, column=col).fill = fill_sect

# ================================================================ CONFIG
cfg = wb.active
cfg.title = "Config"
cfg.sheet_properties.tabColor = "808080"
cfg["A1"] = "CONFIG — every parameter of the model lives here (nothing is hardcoded in formulas)"
cfg["A1"].font = f_title; cfg["A1"].fill = fill_title
for col in range(2, 7): cfg.cell(row=1, column=col).fill = fill_title

sect(cfg, 3, "1. CORE PARAMETERS", 6)
params = [
    # row, label, value, fmt, note, defined name
    (4,  "Reference / reporting currency", "USD", None,
        "All values converted to this currency for margin maths.", "RefCcy"),
    (5,  "Cash asset-type label", "Cash", None,
        "Rows whose Asset Type equals this label are treated as cash: LV 100%, exempt from concentration.", "CashLabel"),
    (6,  "Single Counter concentration threshold", 0.30, PCT,
        "DBS example: 30%. Adj Factor = threshold / ratio when breached.", "ThrSingle"),
    (7,  "Issuer concentration threshold", 0.40, PCT,
        "DBS example: 40%.", "ThrIssuer"),
    (8,  "Country-of-risk threshold (default)", 1.00, PCT,
        "Used when the country is not in table 3 below (DBS example: most = 100%, Macao 75%).", "DefThrCountry"),
    (9,  "Asset-type threshold (default)", 1.00, PCT,
        "Used when no per-asset-type threshold is set in table 2.", "DefThrAsset"),
    (10, "Cash cap in concentration denominator", 0.30, PCT,
        "Denominator = Marketable CV + MIN(Cash CV, this % x Marketable CV) — per DBS PLV report.", "CashCapPct"),
    (11, "HCFX currency-mismatch haircut", 0.10, PCT,
        "10% haircut on CV allocated to exposure in a different currency bucket.", "HCFXRate"),
    (12, "LV cap (marketable securities)", 0.75, PCT,
        "Regulatory LV cap (e.g. 75% cap in HK for equities/ELNs). Applied to non-cash rows.", "LVCap"),
    (13, "Margin call trigger (utilisation of net CV)", 1.00, PCT,
        "Margin call when Exposure > Net CV x this trigger.", "MCTrigger"),
    (14, "Early-warning utilisation level", 0.90, PCT,
        "Status shows 'Warning' above this utilisation.", "WarnLevel"),
]
for r, lbl, val, fmt, note, nm in params:
    cfg.cell(row=r, column=1, value=lbl).font = f_lbl
    c = cfg.cell(row=r, column=2, value=val)
    c.fill = fill_in; c.border = border; c.font = f_lblb
    if fmt: c.number_format = fmt
    cfg.cell(row=r, column=3, value=note).font = f_small
    cfg.cell(row=r, column=3).alignment = wrap
    name(nm, f"Config!$B${r}")

sect(cfg, 17, "2. ASSET TYPES — default Lending Value % and concentration threshold", 6)
hdr_row(cfg, 18, ["Asset Type", "Default LV %", "Conc. Threshold %", "Notes"])
asset_rows = [
    ("Listed Equity",      0.70, 1.00, "LV capped by parameter 'LV cap' above"),
    ("Bond",               0.75, 1.00, ""),
    ("Fund / ETF",         0.65, 1.00, ""),
    ("Structured Product", 0.60, 1.00, "e.g. FCN / ELN"),
    ("Cash",               1.00, 1.00, "Exempt from concentration; HCFX still applies"),
]
LV_FIRST, LV_LAST = 19, 28
for i in range(LV_FIRST, LV_LAST + 1):
    data = asset_rows[i - LV_FIRST] if i - LV_FIRST < len(asset_rows) else ("", None, None, "")
    cfg.cell(row=i, column=1, value=data[0])
    cfg.cell(row=i, column=2, value=data[1]).number_format = PCT
    cfg.cell(row=i, column=3, value=data[2]).number_format = PCT
    cfg.cell(row=i, column=4, value=data[3]).font = f_small
    for col in range(1, 4):
        cfg.cell(row=i, column=col).fill = fill_in
        cfg.cell(row=i, column=col).border = border
name("LVTable", f"Config!$A${LV_FIRST}:$C${LV_LAST}")
name("AssetTypeList", f"Config!$A${LV_FIRST}:$A${LV_LAST}")

sect(cfg, 31, "3. COUNTRY-OF-RISK THRESHOLD OVERRIDES (all other countries use the default in section 1)", 6)
hdr_row(cfg, 32, ["Country", "Threshold %"])
CT_FIRST, CT_LAST = 33, 40
country_rows = [("Macao", 0.75)]
for i in range(CT_FIRST, CT_LAST + 1):
    data = country_rows[i - CT_FIRST] if i - CT_FIRST < len(country_rows) else ("", None)
    cfg.cell(row=i, column=1, value=data[0])
    cfg.cell(row=i, column=2, value=data[1]).number_format = PCT
    for col in range(1, 3):
        cfg.cell(row=i, column=col).fill = fill_in
        cfg.cell(row=i, column=col).border = border
name("CountryThr", f"Config!$A${CT_FIRST}:$B${CT_LAST}")

sect(cfg, 43, "4. FX RATES & HCFX CURRENCY BUCKETS  (rate = reference-ccy per 1 unit; live rate overrides manual when present)", 6)
hdr_row(cfg, 44, ["Currency", "Manual rate", "Live rate (optional)", "Rate used", "HCFX Bucket", "Note"])
FX_FIRST, FX_LAST = 45, 56
fx_rows = [
    ("USD", 1.0000, "USD&HKD", "Reference currency"),
    ("HKD", 0.1282, "USD&HKD", "USD/HKD pegged — one bucket, no HCFX between them"),
    ("CHF", 1.2349, "CHF", ""),
    ("EUR", 1.1700, "EUR", ""),
    ("GBP", 1.3400, "GBP", ""),
    ("SGD", 0.7800, "SGD", ""),
    ("JPY", 0.0067, "JPY", ""),
    ("CNH", 0.1390, "CNH", ""),
]
for i in range(FX_FIRST, FX_LAST + 1):
    k = i - FX_FIRST
    ccy, rate, bucket, note = fx_rows[k] if k < len(fx_rows) else ("", None, "", "")
    cfg.cell(row=i, column=1, value=ccy)
    cfg.cell(row=i, column=2, value=rate).number_format = '0.0000'
    cfg.cell(row=i, column=3).number_format = '0.0000'
    d = cfg.cell(row=i, column=4,
                 value=f'=IF($A{i}="","",IF(ISNUMBER($C{i}),$C{i},$B{i}))')
    d.number_format = '0.0000'; d.fill = fill_calc
    cfg.cell(row=i, column=5, value=bucket)
    cfg.cell(row=i, column=6, value=note).font = f_small
    for col in (1, 2, 3, 5):
        cfg.cell(row=i, column=col).fill = fill_in
    for col in range(1, 6):
        cfg.cell(row=i, column=col).border = border
name("FXRates", f"Config!$A${FX_FIRST}:$E${FX_LAST}")
name("CcyList", f"Config!$A${FX_FIRST}:$A${FX_LAST}")

sect(cfg, 59, "5. HCFX BUCKET LIST (drives the HCFX sheet — add a row here when you add a new bucket in section 4)", 6)
hdr_row(cfg, 60, ["Bucket"])
BK_FIRST, BK_LAST = 61, 70
buckets = ["USD&HKD", "CHF", "EUR", "GBP", "SGD", "JPY", "CNH"]
for i in range(BK_FIRST, BK_LAST + 1):
    k = i - BK_FIRST
    cfg.cell(row=i, column=1, value=buckets[k] if k < len(buckets) else "")
    cfg.cell(row=i, column=1).fill = fill_in
    cfg.cell(row=i, column=1).border = border

sect(cfg, 73, "6. FACILITIES / EXPOSURE (loans drawn — any currency, any number of lines)", 6)
hdr_row(cfg, 74, ["Facility", "Ccy", "Amount (facility ccy)", "Rate used", "Amount (ref ccy)", "HCFX Bucket"])
FAC_FIRST, FAC_LAST = 75, 82
fac_rows = [("Lombard loan", "CHF", 1604693)]
for i in range(FAC_FIRST, FAC_LAST + 1):
    k = i - FAC_FIRST
    namev, ccy, amt = fac_rows[k] if k < len(fac_rows) else ("", "", None)
    cfg.cell(row=i, column=1, value=namev)
    cfg.cell(row=i, column=2, value=ccy)
    cfg.cell(row=i, column=3, value=amt).number_format = MONEY
    r = cfg.cell(row=i, column=4, value=f'=IF($B{i}="","",IFERROR(VLOOKUP($B{i},FXRates,4,FALSE),1))')
    r.number_format = '0.0000'; r.fill = fill_calc
    a = cfg.cell(row=i, column=5, value=f'=IF($B{i}="","",$C{i}*$D{i})')
    a.number_format = MONEY; a.fill = fill_calc
    b = cfg.cell(row=i, column=6, value=f'=IF($B{i}="","",IFERROR(VLOOKUP($B{i},FXRates,5,FALSE),$B{i}))')
    b.fill = fill_calc
    for col in (1, 2, 3):
        cfg.cell(row=i, column=col).fill = fill_in
    for col in range(1, 7):
        cfg.cell(row=i, column=col).border = border
TOTF = FAC_LAST + 1
cfg.cell(row=TOTF, column=1, value="TOTAL EXPOSURE").font = f_lblb
cfg.cell(row=TOTF, column=1).fill = fill_tot
e = cfg.cell(row=TOTF, column=5, value=f"=SUM(E{FAC_FIRST}:E{FAC_LAST})")
e.number_format = MONEY; e.font = f_lblb; e.fill = fill_tot; e.border = border
name("ExposureTot", f"Config!$E${TOTF}")
name("FacAmt", f"Config!$E${FAC_FIRST}:$E${FAC_LAST}")
name("FacBucket", f"Config!$F${FAC_FIRST}:$F${FAC_LAST}")

for col, w in zip("ABCDEF", [42, 16, 22, 12, 16, 42]):
    cfg.column_dimensions[col].width = w
cfg.cell(row=TOTF + 2, column=1,
         value="Yellow cells = inputs. Blue cells = formulas. Add rows inside the marked table ranges only.").font = f_small

# ================================================================ POSITIONS
pos = wb.create_sheet("Positions")
pos.sheet_properties.tabColor = "2E5AA8"
pos["A1"] = "POSITIONS & COLLATERAL ENGINE — paste any portfolio here; everything below row 4 recalculates"
pos["A1"].font = f_title; pos["A1"].fill = fill_title
for col in range(2, 38): pos.cell(row=1, column=col).fill = fill_title

# engine summary block (row 2) ------------------------------------
CL = "CashLabel"
eng = [
    ("B", "Marketable CV:",       f'=SUMIFS($P${R0}:$P${R1},$D${R0}:$D${R1},"<>"&{CL},$B${R0}:$B${R1},"<>")'),
    ("D", "Cash CV:",             f'=SUMIFS($P${R0}:$P${R1},$D${R0}:$D${R1},{CL})'),
    ("F", "Cash recognised:",     '=MIN($E$2,$C$2*CashCapPct)'),
    ("H", "Conc denominator:",    '=$C$2+$G$2'),
    ("J", "Str. marketable CV:",  f'=SUMIFS($AE${R0}:$AE${R1},$D${R0}:$D${R1},"<>"&{CL},$B${R0}:$B${R1},"<>")'),
    ("L", "Str. cash CV:",        f'=SUMIFS($AE${R0}:$AE${R1},$D${R0}:$D${R1},{CL})'),
    ("N", "Str. cash recognised:", '=MIN($M$2,$K$2*CashCapPct)'),
    ("P", "Str. denominator:",    '=$K$2+$O$2'),
]
for colL, lbl, f in eng:
    ci = openpyxl.utils.column_index_from_string(colL)
    pos.cell(row=2, column=ci, value=lbl).font = f_small
    v = pos.cell(row=2, column=ci + 1, value=f)
    v.number_format = MONEY; v.font = Font(bold=True, size=9); v.fill = fill_calc
name("ConcDenom", "Positions!$I$2")
name("ConcDenomStr", "Positions!$Q$2")
pos.cell(row=3, column=2,
         value="Denominator = Marketable CV + MIN(Cash CV, cap% x Marketable CV)   |   cash rows: LV 100%, concentration-exempt, HCFX still applies").font = f_small

headers = [
    "Line", "Security / Position", "Ticker", "Asset Type", "Issuer", "Country of Risk",
    "Ccy", "Quantity", "Price (manual)", "Price (live, optional)", "Price used",
    "FX rate used", "Market Value (ref ccy)", "LV% override", "LV% used",
    "Collateral Value (CV)", "HCFX Bucket",
    "Single-Counter ratio", "Issuer ratio", "Country ratio", "Asset-Type ratio",
    "Adj Factor Single", "Adj Factor Issuer", "Adj Factor Country", "Adj Factor Asset",
    "Min Adj Factor", "Conc Adj CV", "Concentration Penalty",
    "Price shock % (input)", "Stressed MV", "Stressed CV",
    "s-Single ratio", "s-Issuer ratio", "s-Country ratio", "s-Asset ratio",
    "s-Min Factor", "Stressed Conc Adj CV",
]
hdr_row(pos, 4, headers)
pos.row_dimensions[4].height = 42

CTHR = '=IFERROR(VLOOKUP($F{r},CountryThr,2,FALSE),DefThrCountry)'.lstrip("=")
ATHR = '=IFERROR(VLOOKUP($D{r},LVTable,3,FALSE),DefThrAsset)'.lstrip("=")

def ratio(sumcol, keycol, r, denom):
    """conc ratio of this row's group in dimension keyed by keycol, over CV column sumcol"""
    return (f'=IF($B{r}="","",IF($D{r}={CL},"",'
            f'SUMIFS(${sumcol}${R0}:${sumcol}${R1},${keycol}${R0}:${keycol}${R1},${keycol}{r},'
            f'$D${R0}:$D${R1},"<>"&{CL})/{denom}))')

for r in range(R0, R1 + 1):
    F = {
        'A': f'=IF($B{r}="","",ROW()-{R0 - 1})',
        'K': f'=IF($B{r}="","",IF(ISNUMBER($J{r}),$J{r},$I{r}))',
        'L': f'=IF($B{r}="","",IFERROR(VLOOKUP($G{r},FXRates,4,FALSE),1))',
        'M': f'=IF($B{r}="","",$H{r}*$K{r}*$L{r})',
        'O': (f'=IF($B{r}="","",IF($D{r}={CL},IF(ISNUMBER($N{r}),$N{r},1),'
              f'MIN(IF(ISNUMBER($N{r}),$N{r},IFERROR(VLOOKUP($D{r},LVTable,2,FALSE),0)),LVCap)))'),
        'P': f'=IF($B{r}="","",$M{r}*$O{r})',
        'Q': f'=IF($B{r}="","",IFERROR(VLOOKUP($G{r},FXRates,5,FALSE),$G{r}))',
        'R': ratio('P', 'B', r, 'ConcDenom'),
        'S': ratio('P', 'E', r, 'ConcDenom'),
        'T': ratio('P', 'F', r, 'ConcDenom'),
        'U': ratio('P', 'D', r, 'ConcDenom'),
        'V': f'=IF($B{r}="","",IF($D{r}={CL},1,IF($R{r}>ThrSingle,ThrSingle/$R{r},1)))',
        'W': f'=IF($B{r}="","",IF($D{r}={CL},1,IF($S{r}>ThrIssuer,ThrIssuer/$S{r},1)))',
        'X': (f'=IF($B{r}="","",IF($D{r}={CL},1,'
              f'IF($T{r}>{CTHR.format(r=r)},{CTHR.format(r=r)}/$T{r},1)))'),
        'Y': (f'=IF($B{r}="","",IF($D{r}={CL},1,'
              f'IF($U{r}>{ATHR.format(r=r)},{ATHR.format(r=r)}/$U{r},1)))'),
        'Z': f'=IF($B{r}="","",IF($D{r}={CL},1,MIN($V{r}:$Y{r})))',
        'AA': f'=IF($B{r}="","",$P{r}*$Z{r})',
        'AB': f'=IF($B{r}="","",$P{r}-$AA{r})',
        'AD': f'=IF($B{r}="","",$M{r}*(1+IF(ISNUMBER($AC{r}),$AC{r},0)))',
        'AE': f'=IF($B{r}="","",$AD{r}*$O{r})',
        'AF': ratio('AE', 'B', r, 'ConcDenomStr'),
        'AG': ratio('AE', 'E', r, 'ConcDenomStr'),
        'AH': ratio('AE', 'F', r, 'ConcDenomStr'),
        'AI': ratio('AE', 'D', r, 'ConcDenomStr'),
        'AJ': (f'=IF($B{r}="","",IF($D{r}={CL},1,MIN('
               f'IF($AF{r}>ThrSingle,ThrSingle/$AF{r},1),'
               f'IF($AG{r}>ThrIssuer,ThrIssuer/$AG{r},1),'
               f'IF($AH{r}>{CTHR.format(r=r)},{CTHR.format(r=r)}/$AH{r},1),'
               f'IF($AI{r}>{ATHR.format(r=r)},{ATHR.format(r=r)}/$AI{r},1))))'),
        'AK': f'=IF($B{r}="","",$AE{r}*$AJ{r})',
    }
    for colL, f in F.items():
        pos[f"{colL}{r}"] = f

# sample portfolio (from the client's uploaded monitor)
sample = [
    # name, ticker, asset type, issuer, country, ccy, qty, price
    ("Meta Platforms Inc", "META", "Listed Equity", "Meta Platforms Inc", "United States", "USD", 2800, 542.87),
    ("Roblox Corp", "RBLX", "Listed Equity", "Roblox Corp", "United States", "USD", 9800, 46.39),
    ("FCN on HOOD (Robinhood)", "", "Structured Product", "Issuing Bank A", "United States", "USD", 1, 600000),
    ("FCN on RBLX (Roblox)", "", "Structured Product", "Issuing Bank A", "United States", "USD", 1, 600000),
    ("Time deposit USD", "", "Cash", "Deposit Bank", "Hong Kong", "USD", 1, 708197.43),
    ("Time deposit HKD", "", "Cash", "Deposit Bank", "Hong Kong", "HKD", 1, 824747.59),
    ("Fenghe Asia Fund (USTE)", "", "Fund / ETF", "Fenghe Asset Management", "Singapore", "USD", 1, 1180.82),
]
demo_shocks = [-0.25, -0.25, -0.15, -0.15, 0, 0, -0.10]   # demo stress scenario, editable
for i, row in enumerate(sample):
    r = R0 + i
    for j, v in enumerate(row):
        pos.cell(row=r, column=2 + j, value=v)
    pos.cell(row=r, column=29, value=demo_shocks[i])       # col AC price shock

# totals row
pos.cell(row=TOT, column=1, value="TOTAL").font = f_lblb
for colL in ["M", "P", "AA", "AB", "AD", "AE", "AK"]:
    c = pos[f"{colL}{TOT}"]
    c.value = f"=SUM({colL}{R0}:{colL}{R1})"
    c.font = f_lblb
for col in range(1, 38):
    pos.cell(row=TOT, column=col).fill = fill_tot
    pos.cell(row=TOT, column=col).border = border

# formats / fills
fmt_map = {
    'H': MONEY2, 'I': MONEY2, 'J': MONEY2, 'K': MONEY2, 'L': '0.0000',
    'M': MONEY, 'N': PCT, 'O': PCT, 'P': MONEY,
    'R': PCT, 'S': PCT, 'T': PCT, 'U': PCT,
    'V': FACT, 'W': FACT, 'X': FACT, 'Y': FACT, 'Z': FACT,
    'AA': MONEY, 'AB': MONEY, 'AC': PCT, 'AD': MONEY, 'AE': MONEY,
    'AF': PCT, 'AG': PCT, 'AH': PCT, 'AI': PCT, 'AJ': FACT, 'AK': MONEY,
}
input_cols = {'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'N', 'AC'}
for r in range(R0, R1 + 1):
    for colL, fmt in fmt_map.items():
        pos[f"{colL}{r}"].number_format = fmt
    for colL in input_cols:
        pos[f"{colL}{r}"].fill = fill_in
    for colL in ('P', 'AA', 'Z'):
        pos[f"{colL}{r}"].fill = fill_calc
    for col in range(1, 38):
        pos.cell(row=r, column=col).border = border
for colL in ["M", "P", "AA", "AB", "AD", "AE", "AK"]:
    pos[f"{colL}{TOT}"].number_format = MONEY

widths = {'A': 5, 'B': 26, 'C': 8, 'D': 16, 'E': 22, 'F': 15, 'G': 6, 'H': 10,
          'I': 11, 'J': 11, 'K': 11, 'L': 9, 'M': 14, 'N': 9, 'O': 8, 'P': 14,
          'Q': 10, 'R': 9, 'S': 9, 'T': 9, 'U': 9, 'V': 8, 'W': 8, 'X': 8,
          'Y': 8, 'Z': 8, 'AA': 14, 'AB': 13, 'AC': 9, 'AD': 13, 'AE': 13,
          'AF': 9, 'AG': 9, 'AH': 9, 'AI': 9, 'AJ': 8, 'AK': 14}
for colL, w in widths.items():
    pos.column_dimensions[colL].width = w
pos.freeze_panes = "C5"

# data validation
dv_asset = DataValidation(type="list", formula1="=AssetTypeList", allow_blank=True,
                          showErrorMessage=False)
dv_ccy = DataValidation(type="list", formula1="=CcyList", allow_blank=True,
                        showErrorMessage=False)
pos.add_data_validation(dv_asset); pos.add_data_validation(dv_ccy)
dv_asset.add(f"D{R0}:D{R1}")
dv_ccy.add(f"G{R0}:G{R1}")

# conditional formatting: factors < 1 amber; penalty > 0 red
amber = PatternFill("solid", fgColor=WARN)
red = PatternFill("solid", fgColor=BAD)
pos.conditional_formatting.add(
    f"V{R0}:Z{R1}",
    CellIsRule(operator="lessThan", formula=["1"], fill=amber, font=Font(color=WARNF)))
pos.conditional_formatting.add(
    f"AJ{R0}:AJ{R1}",
    CellIsRule(operator="lessThan", formula=["1"], fill=amber, font=Font(color=WARNF)))
pos.conditional_formatting.add(
    f"AB{R0}:AB{R1}",
    CellIsRule(operator="greaterThan", formula=["0.005"], fill=red, font=Font(color=BADF)))

# ================================================================ HCFX
hc = wb.create_sheet("HCFX")
hc.sheet_properties.tabColor = "C55A11"
hc["A1"] = "HCFX — CURRENCY MISMATCH HAIRCUT (buckets & rate come from Config; USD & HKD = one bucket)"
hc["A1"].font = f_title; hc["A1"].fill = fill_title
for col in range(2, 11): hc.cell(row=1, column=col).fill = fill_title
hc["A2"] = ("HCFX = (CV − non-allocated CV − CV allocated to same-currency exposure) x haircut rate. "
            "Conservative treatment per the training example: when any liability is unmet in its own bucket, "
            "the full residual CV of every bucket is haircut.")
hc["A2"].font = f_small; hc["A2"].alignment = wrap
hc.merge_cells("A2:J2"); hc.row_dimensions[2].height = 26

hdr_row(hc, 4, ["Bucket", "Liabilities (ref ccy)", "CV Used (Conc Adj)", "Residual CV",
                "Unmet Liability", "Currency Haircut",
                "CV Used (stressed)", "Residual CV (str)", "Unmet Liab (str)", "Haircut (str)"])
hc.row_dimensions[4].height = 30
HC_FIRST = 5
HC_LAST = HC_FIRST + (BK_LAST - BK_FIRST)   # one row per Config bucket row
HCT = HC_LAST + 1
for i in range(HC_FIRST, HC_LAST + 1):
    src = BK_FIRST + (i - HC_FIRST)
    hc[f"A{i}"] = f'=IF(Config!$A${src}="","",Config!$A${src})'
    hc[f"B{i}"] = f'=IF($A{i}="","",SUMIFS(FacAmt,FacBucket,$A{i}))'
    hc[f"C{i}"] = f'=IF($A{i}="","",SUMIFS(Positions!$AA${R0}:$AA${R1},Positions!$Q${R0}:$Q${R1},$A{i}))'
    hc[f"D{i}"] = f'=IF($A{i}="","",MAX(0,$C{i}-$B{i}))'
    hc[f"E{i}"] = f'=IF($A{i}="","",MAX(0,$B{i}-$C{i}))'
    hc[f"F{i}"] = f'=IF($A{i}="","",$D{i}*HCFXRate*IF($E${HCT}>0,1,0))'
    hc[f"G{i}"] = f'=IF($A{i}="","",SUMIFS(Positions!$AK${R0}:$AK${R1},Positions!$Q${R0}:$Q${R1},$A{i}))'
    hc[f"H{i}"] = f'=IF($A{i}="","",MAX(0,$G{i}-$B{i}))'
    hc[f"I{i}"] = f'=IF($A{i}="","",MAX(0,$B{i}-$G{i}))'
    hc[f"J{i}"] = f'=IF($A{i}="","",$H{i}*HCFXRate*IF($I${HCT}>0,1,0))'
    for colL in "BCDEFGHIJ":
        hc[f"{colL}{i}"].number_format = MONEY
    for col in range(1, 11):
        hc.cell(row=i, column=col).border = border
hc.cell(row=HCT, column=1, value="TOTAL").font = f_lblb
for colL in "BCDEFGHIJ":
    c = hc[f"{colL}{HCT}"]
    c.value = f"=SUM({colL}{HC_FIRST}:{colL}{HC_LAST})"
    c.number_format = MONEY; c.font = f_lblb
for col in range(1, 11):
    hc.cell(row=HCT, column=col).fill = fill_tot
    hc.cell(row=HCT, column=col).border = border
name("HCFXTot", f"HCFX!$F${HCT}")
name("HCFXTotStr", f"HCFX!$J${HCT}")
hc.cell(row=HCT + 2, column=1,
        value="CV Used is taken AFTER the concentration haircut (Conc Adj CV), so the waterfall is: "
              "CV -> concentration haircut -> HCFX -> Net CV.").font = f_small
for colL, w in zip("ABCDEFGHIJ", [12, 16, 16, 14, 14, 14, 16, 14, 14, 14]):
    hc.column_dimensions[colL].width = w

# ================================================================ DASHBOARD
db = wb.create_sheet("Dashboard", 0)
db.sheet_properties.tabColor = "1F3864"
db["A1"] = "SECURITIES-BACKED LENDING — MARGIN CALL DASHBOARD"
db["A1"].font = f_title; db["A1"].fill = fill_title
for col in range(2, 8): db.cell(row=1, column=col).fill = fill_title
db["A2"] = '="Reference currency: "&RefCcy&"   |   As of: "&TEXT(TODAY(),"yyyy-mm-dd")&"   |   Methodology: CV = Qty x Price x LV%  ->  Concentration haircut (4 dimensions)  ->  HCFX  ->  Net CV"'
db["A2"].font = f_small

db["A4"] = "CURRENT"; db["A4"].font = f_sect; db["A4"].fill = fill_sect
db["B4"].fill = fill_sect
db["D4"] = "STRESSED (per-position shocks on Positions col AC)"; db["D4"].font = f_sect; db["D4"].fill = fill_sect
db["E4"].fill = fill_sect

cur = [
    ("Total exposure (all facilities)", "=ExposureTot", MONEY),
    ("Gross collateral market value", f"=Positions!$M${TOT}", MONEY),
    ("Collateral Value before haircuts (CV)", f"=Positions!$P${TOT}", MONEY),
    ("Concentration penalty", f"=-Positions!$AB${TOT}", MONEY),
    ("Conc Adj CV (after concentration haircut)", f"=Positions!$AA${TOT}", MONEY),
    ("Currency mismatch haircut (HCFX)", "=-HCFXTot", MONEY),
    ("NET COLLATERAL VALUE (CV2)", f"=Positions!$AA${TOT}-HCFXTot", MONEY),
    ("Utilisation (Exposure / Net CV)", "=IFERROR(ExposureTot/NetCV,0)", PCT2),
    ("STATUS", '=IF(NetCV<=0,"MARGIN CALL",IF(ExposureTot/NetCV>=MCTrigger,"MARGIN CALL",IF(ExposureTot/NetCV>=WarnLevel,"WARNING","OK")))', None),
    ("Buffer to margin call", "=NetCV*MCTrigger-ExposureTot", MONEY),
    ("Top-up required now", "=MAX(0,ExposureTot-NetCV*MCTrigger)", MONEY),
    ("Collateral fall to WARNING (uniform, approx)", "=IF(NetCV=0,0,MAX(0,1-ExposureTot/(WarnLevel*NetCV)))", PCT),
    ("Collateral fall to MARGIN CALL (uniform, approx)", "=IF(NetCV=0,0,MAX(0,1-ExposureTot/(MCTrigger*NetCV)))", PCT),
    ("Required CV incl. FX buffer  [Exposure / (1-HCFX)]", "=ExposureTot/(1-HCFXRate)", MONEY),
]
strs = [
    ("Weighted average price shock", f"=IFERROR(Positions!$AD${TOT}/Positions!$M${TOT}-1,0)", PCT),
    ("Stressed gross market value", f"=Positions!$AD${TOT}", MONEY),
    ("Stressed CV before haircuts", f"=Positions!$AE${TOT}", MONEY),
    ("Stressed concentration penalty", f"=-(Positions!$AE${TOT}-Positions!$AK${TOT})", MONEY),
    ("Stressed Conc Adj CV", f"=Positions!$AK${TOT}", MONEY),
    ("Stressed HCFX", "=-HCFXTotStr", MONEY),
    ("STRESSED NET COLLATERAL VALUE", f"=Positions!$AK${TOT}-HCFXTotStr", MONEY),
    ("Stressed utilisation", "=IFERROR(ExposureTot/NetCVStr,0)", PCT2),
    ("STRESSED STATUS", '=IF(NetCVStr<=0,"MARGIN CALL",IF(ExposureTot/NetCVStr>=MCTrigger,"MARGIN CALL",IF(ExposureTot/NetCVStr>=WarnLevel,"WARNING","OK")))', None),
    ("Stressed buffer to margin call", "=NetCVStr*MCTrigger-ExposureTot", MONEY),
    ("Stressed top-up required", "=MAX(0,ExposureTot-NetCVStr*MCTrigger)", MONEY),
]
for i, (lbl, f, fmt) in enumerate(cur):
    r = 5 + i
    db.cell(row=r, column=1, value=lbl).font = f_lbl
    c = db.cell(row=r, column=2, value=f)
    c.font = f_kpi; c.border = border; c.fill = fill_grey
    if fmt: c.number_format = fmt
for i, (lbl, f, fmt) in enumerate(strs):
    r = 5 + i
    db.cell(row=r, column=4, value=lbl).font = f_lbl
    c = db.cell(row=r, column=5, value=f)
    c.font = f_kpi; c.border = border; c.fill = fill_grey
    if fmt: c.number_format = fmt
name("NetCV", "Dashboard!$B$11")
name("NetCVStr", "Dashboard!$E$11")
for cell in ("A11", "B11", "D11", "E11", "A13", "B13", "D13", "E13"):
    db[cell].font = Font(bold=True, size=11)

# concentration monitor
row = 21
db.cell(row=row, column=1, value="CONCENTRATION MONITOR (worst group per dimension)").font = f_sect
for col in range(1, 8): db.cell(row=row, column=col).fill = fill_sect
hdr_row(db, row + 1, ["Dimension", "Threshold", "Max ratio", "Min Adj Factor", "Status", "Driver (worst position row)"])
dims = [
    ("Single Counter", "=ThrSingle", "R", "V"),
    ("Issuer",         "=ThrIssuer", "S", "W"),
    ("Country of Risk", '="default "&TEXT(DefThrCountry,"0%")&" / overrides"', "T", "X"),
    ("Asset Type",      '="default "&TEXT(DefThrAsset,"0%")&" / per type"', "U", "Y"),
]
for i, (dim, thr, rc, fc) in enumerate(dims):
    r = row + 2 + i
    db.cell(row=r, column=1, value=dim).font = f_lbl
    t = db.cell(row=r, column=2, value=thr)
    if not thr.startswith('="'): t.number_format = PCT
    m = db.cell(row=r, column=3, value=f"=IF(COUNT(Positions!${rc}${R0}:${rc}${R1})=0,0,MAX(Positions!${rc}${R0}:${rc}${R1}))")
    m.number_format = PCT
    mf = db.cell(row=r, column=4, value=f"=IF(COUNT(Positions!${fc}${R0}:${fc}${R1})=0,1,MIN(Positions!${fc}${R0}:${fc}${R1}))")
    mf.number_format = FACT
    db.cell(row=r, column=5, value=f'=IF($D{r}<1,"BREACH","OK")')
    db.cell(row=r, column=6, value=(f'=IF($D{r}<1,INDEX(Positions!$B${R0}:$B${R1},'
                                    f'MATCH($D{r},Positions!${fc}${R0}:${fc}${R1},0)),"—")')).font = f_lbl
    for col in range(1, 7):
        db.cell(row=r, column=col).border = border

row = row + 7
db.cell(row=row, column=1, value="HOW TO READ").font = f_sect
for col in range(1, 8): db.cell(row=row, column=col).fill = fill_sect
notes = [
    "1. CV per position = Quantity x Price x LV% (LV defaults by asset type on Config, capped by the LV cap; override per line).",
    "2. Concentration: ratio of each group's CV over [Marketable CV + MIN(Cash, cap% x Marketable)]. If a ratio breaches its threshold, Adj Factor = threshold / ratio; Conc Adj CV = CV x MIN of the 4 factors (conservative: the single most-penal dimension binds).",
    "3. HCFX: collateral supporting exposure in another currency bucket is haircut (default 10%); USD & HKD share one bucket. See HCFX sheet.",
    "4. Margin call when Exposure > Net CV x trigger. All thresholds live on Config — nothing is hardcoded.",
    "5. Stress: enter price shocks per line in Positions col AC (e.g. -20%). The whole chain (concentration + HCFX) is recomputed under stress.",
]
for i, n in enumerate(notes):
    db.cell(row=row + 1 + i, column=1, value=n).font = f_small
for colL, w in zip("ABCDEFG", [44, 18, 12, 13, 12, 30, 12]):
    db.column_dimensions[colL].width = w

# status conditional formatting
for rng in ("B13", "E13", "E23:E26"):
    db.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("MARGIN CALL",{rng.split(":")[0]}))'],
        fill=PatternFill("solid", fgColor=BAD), font=Font(color=BADF, bold=True)))
    db.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("WARNING",{rng.split(":")[0]}))'],
        fill=PatternFill("solid", fgColor=WARN), font=Font(color=WARNF, bold=True)))
    db.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("BREACH",{rng.split(":")[0]}))'],
        fill=PatternFill("solid", fgColor=BAD), font=Font(color=BADF, bold=True)))
    db.conditional_formatting.add(rng, FormulaRule(
        formula=[f'{rng.split(":")[0]}="OK"'],
        fill=PatternFill("solid", fgColor=GOOD), font=Font(color=GOODF, bold=True)))

# ================================================================ LIVE DATA GUIDE
gd = wb.create_sheet("Live Data Guide")
gd.sheet_properties.tabColor = "70AD47"
gd["A1"] = "LIVE PRICES & FX — OPTIONS AND EXACT WIRING"
gd["A1"].font = f_title; gd["A1"].fill = fill_title
for col in range(2, 4): gd.cell(row=1, column=col).fill = fill_title
guide = [
    ("", ""),
    ("HOW THE WORKBOOK CONSUMES LIVE DATA", ""),
    ("Prices", "Positions col J 'Price (live, optional)': if it contains a number it overrides the manual price in col I automatically (col K = IF(ISNUMBER(J),J,I)). Point col J at any of the sources below."),
    ("FX rates", "Config section 4 col C 'Live rate (optional)': same override logic — col D uses the live rate when present, else the manual rate."),
    ("", ""),
    ("OPTION 1 — EXCEL 365 STOCKS & CURRENCIES DATA TYPES (recommended, no add-in)", ""),
    ("Prices", "Type the ticker in a helper cell (e.g. col C), select it, Data tab -> Data Types -> Stocks. Then in col J:  =FIELDVALUE($C5,\"Price\").  Refresh: Data -> Refresh All (or right-click -> Data Type -> Refresh). Quotes are delayed ~15 min."),
    ("FX", "Type the pair (e.g.  CHF/USD ) in a helper cell, convert to Currencies data type, then:  =FIELDVALUE(cell,\"Price\")  in Config col C."),
    ("", ""),
    ("OPTION 2 — STOCKHISTORY (Excel 365, latest close)", ""),
    ("Prices", "Formula:  =IFERROR(INDEX(STOCKHISTORY($C5,TODAY()-7,TODAY(),0,0,1),ROWS(STOCKHISTORY($C5,TODAY()-7,TODAY(),0,0,1)),2),\"\")   -> returns the most recent close for the ticker in C5."),
    ("FX", "Formula:  =INDEX(STOCKHISTORY(\"CHFUSD\",TODAY()-7,TODAY(),0,0,1),ROWS(STOCKHISTORY(\"CHFUSD\",TODAY()-7,TODAY(),0,0,1)),2)"),
    ("", ""),
    ("OPTION 3 — POWER QUERY FROM A FREE API (any Excel with Get & Transform)", ""),
    ("FX", "Data -> Get Data -> From Web ->  https://api.frankfurter.app/latest?from=USD  (ECB rates, free, no key). Load to a table, set 'Refresh on open' + background refresh in the connection properties, and point Config col C at the loaded table with VLOOKUP/INDEX."),
    ("Prices", "Same approach against your data vendor's REST endpoint (e.g. stooq.com CSV: https://stooq.com/q/l/?s=meta.us&f=sd2t2ohlcv&h&e=csv)."),
    ("", ""),
    ("OPTION 4 — TERMINAL / RTD ADD-INS (true real-time, needs entitlement)", ""),
    ("Bloomberg", 'Formulas:  =BDP("META US Equity","PX_LAST")     and     =BDP("USDCHF Curncy","PX_LAST")'),
    ("LSEG/Refinitiv", 'Formulas:  =RtGet("IDN",".META.O","TRDPRC_1")  or  =TR("META.O","TR.PriceClose")'),
    ("IBKR / others", "Interactive Brokers Excel RTD add-in: =RTD(\"ibkrtws.rtd\",,\"META\",\"Last\")"),
    ("", ""),
    ("OPTION 5 — GOOGLE SHEETS BRIDGE (free, near-real-time)", ""),
    ("How", "Maintain a small Google Sheet with =GOOGLEFINANCE(\"META\",\"price\") and =GOOGLEFINANCE(\"CURRENCY:CHFUSD\"), publish it as CSV, then pull it into this workbook via Power Query with auto-refresh."),
    ("", ""),
    ("OPTION 6 — SCRIPTED REFRESH (ops-friendly)", ""),
    ("How", "A scheduled Python job (yfinance / vendor API + openpyxl or xlwings) writes prices into Positions col J and FX into Config col C, leaving all formulas intact. Good for end-of-day batch monitoring."),
    ("", ""),
    ("GOVERNANCE NOTES", ""),
    ("1", "Keep manual price/FX columns populated as fallback — the workbook automatically falls back when a live feed returns an error or blank."),
    ("2", "For credit decisions, snapshot the workbook (values-only copy) at the monitoring cut-off time; live cells keep moving."),
    ("3", "Delayed quotes (~15 min) are fine for margin monitoring; use terminal RTD feeds for intraday close-out decisions."),
]
r = 2
for a, b in guide:
    ca = gd.cell(row=r, column=1, value=a)
    cb = gd.cell(row=r, column=2, value=b)
    if b == "" and a:
        ca.font = f_sect; ca.fill = fill_sect; cb.fill = fill_sect
    else:
        ca.font = f_lblb; cb.font = f_lbl; cb.alignment = wrap
    r += 1
gd.column_dimensions["A"].width = 18
gd.column_dimensions["B"].width = 150
for i in range(2, r):
    gd.row_dimensions[i].height = None

# ================================================================ finish
wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("saved", OUT)
